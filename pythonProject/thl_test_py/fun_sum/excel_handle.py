# -*- coding:utf-8 -*-

import pandas as pd
import openpyxl
from time import sleep
import json


def excel_handle_output_1(
        getdata_fun
        , input_excel_path
        , input_col_name
        , output_excel_path
        , output_col_name
        , output_index
        , temp_result
        , input_sheet_name=''):
    """
    逐个获取本地excel文档的指定列数据，传入到对应的 getdata_fun 函数里；
    然后将返回数据，逐个写入（通过指定output_index，控制对应行）至excel文档（会另建文档）的指定列里。
    :param getdata_fun:获取数据的函数
    :param input_excel_path:输入excel文档的路径
    :param input_col_name:输入excel文档的列名
    :param output_excel_path:输出excel文档的路径
    :param output_col_name:输出excel文档的列名
    :param output_index:由第几行开始输出
    :param temp_result:默认输出值
    :param input_sheet_name:输入excel文档的sheet名
    :return:output_excel_path下的excel文档
    """
    # 取待处理文件
    global col_idx
    # 默认取第一个sheet
    if input_sheet_name == '':
        data_df = pd.read_excel(input_excel_path)
    else:
        data_df = pd.read_excel(input_excel_path, sheet_name=input_sheet_name)
    # 取日期列，赋值到date_list
    date_list = data_df[input_col_name].tolist()
    # 加载excel文档
    wb = openpyxl.load_workbook(input_excel_path)
    # 默认开第一个sheet
    if input_sheet_name == '':
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb[input_sheet_name]
    # 取sheet里的最大列数
    max_col_num = ws.max_column
    # 储存首行列名的list
    col_list = []
    for col in range(1, max_col_num + 1):
        col_list.append(ws.cell(row=1, column=col).value)

    # 通过列名的list，找到output_col_name所在的位置，并转为序号（第几列）
    col_num = col_list.index(output_col_name) + 1
    # 判断最大列数是否超26位，转为excel表的序号
    if (max_col_num > 26) & (col_num > 26):
        col_idx = "%c%c" % (chr(64 + int(col_num / 26)), chr(64 + col_num % 26))
    else:
        col_idx = chr(64 + col_num)

    # 以此遍历地址列表，并按顺序写入到excel
    for cell_data in date_list[(output_index - 2):]:
        try:
            # getdata_fun为调用获取API数据的函数（重点）
            output_data = getdata_fun(str(cell_data))
            ws[col_idx + str(output_index)] = output_data
        except:
            ws[col_idx + str(output_index)] = temp_result
        wb.save(output_excel_path)
        output_index = output_index + 1
        print("序号:" + str(output_index - 2) + "  :  " + str(cell_data) + " —— 导出成功！")
        sleep(2)


def data_json_trans(input_path, sheet_name, output_path):
    """
    将Excel文档转为json文档
    :param input_path: 输入的excel文档路径（要带.xlsx）
    :param sheet_name: 工作表名称
    :param output_path: 输出的json文档路径（要带.json）
    :return: [{列名1：value1,列名2：value2,列名3：value3,...},{列名1：value1,列名2：value2,列名3：value3,...}]
    """
    json_dict_list = []
    data_df = pd.read_excel(input_path, sheet_name=sheet_name)
    # 获取列名
    col_list = data_df.columns.tolist()
    # 获取int64列名
    int64_col_list = data_df.select_dtypes(include=['int64']).columns.tolist()
    # 定义存储“int64列索引”的列表
    int64_col_i_list = []
    # 遍历int64列名列表
    for col_str in int64_col_list:
        # 插入对应索引到int64_col_i_list
        int64_col_i_list.append(col_list.index(col_str))
    # 将每列int64列转为float列
    for int64_col_name in int64_col_list:
        data_df = data_df.astype({int64_col_name: float})
    # 遍历每行
    for col_num in range(data_df.shape[0]):
        json_dict = {}
        # 遍历每列
        for row_num in range(data_df.shape[1]):
            # json_dict = {列名1：value1,列名2：value2,列名3：value3,...}
            # 将int64列的值转为int，其余正常写入
            if row_num in int64_col_i_list:
                json_dict[col_list[row_num]] = int(data_df.iloc[col_num, row_num])
            else:
                json_dict[col_list[row_num]] = data_df.iloc[col_num, row_num]
        # 每行一个字典，插入到json_dict_list列表中
        json_dict_list.append(json_dict)
    print(json_dict_list)
    # 将处理好的json_dict_list写入到输出路径的文档
    with open(output_path, 'w') as f:
        json.dump(json_dict_list, f)
        print(output_path + "————输出成功！")


def json_data_trans(data_json_list, output_dir, output_file_name):
    """
    将json列表输出为excel文档
    :param data_json_list:[{列名1：value1,列名2：value2,列名3：value3,...},{列名1：value1,列名2：value2,列名3：value3,...}]，列表形式
    :param output_dir: 输出excel文档的目录路径
    :param output_file_name: 输出excel文档的名称
    :return: 提示
    """
    data_df = pd.DataFrame(
        data_json_list
    )
    with pd.ExcelWriter(output_dir + "\%s" % output_file_name) as writer:
        data_df.to_excel(writer, sheet_name='data', index=0)
    print("输出为Excel文档成功，文件路径：" + output_dir + "\%s" % output_file_name)
